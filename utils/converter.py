import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

VAT_RATE = 0.18

# Built-in column → (credit_account, is_vat_exempt)
KNOWN_COLUMNS = {
    "שכט":     (5000, False),
    "נוטריון": (5004, False),
    "הוצאות":  (5002, True),
}

INVOICE_HEADERS = [
    "תאריך", "חשבון חובה 1", "חשבון זכות 1", "חשבון זכות 2",
    "פרטים", "אסמכתא", "לקוח-סכום חובה 1", "הכנסה-סכום זכות 1", "מעמע-סכום זכות 2",
]

RECEIPT_HEADERS = [
    "תאריך", "חשבון חובה 1", "חשבון זכות",
    "פרטים", "אסמכתא", "בנק-סכום חובה 1", "לקוח-סכום זכות 2",
]


def detect_month_label(filename: str, df: pd.DataFrame) -> str:
    """Return label like '03.2026' from filename or first date in data."""
    match = re.search(r'[_\-\s](\d{1,2})[_\-\s\.](\d{4})', filename)
    if match:
        m, y = int(match.group(1)), match.group(2)
        return f"{m:02d}.{y}"

    for i in range(1, min(5, len(df))):
        val = df.iloc[i, 1]
        if pd.notna(val) and hasattr(val, "month"):
            return f"{val.month:02d}.{val.year}"

    return "חדש"


def _clean(val) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


def convert_income_file(df: pd.DataFrame, clients_dict: dict, extra_columns: dict):
    """
    Convert income DataFrame to Hashavshevet rows.

    Income file column layout (0-indexed):
      0  סהכ   1  תאריך   2  מעמ    3  סכום
      4  נוטריון  5  הוצאות  6  שכט   7  שונות
      8  תיק   9  לקוח   10  מספר (invoice)

    Returns (invoice_rows, receipt_rows, unmatched_clients, unknown_columns)
    """
    from utils.matcher import flexible_match

    invoice_rows    = []
    receipt_rows    = []
    unmatched       = []
    unknown_columns = set()

    for i in range(1, len(df)):          # skip header row 0
        row = df.iloc[i]

        total       = _clean(row.iloc[0])
        txn_date    = row.iloc[1]
        notary      = _clean(row.iloc[4]) if len(row) > 4 else 0
        expenses    = _clean(row.iloc[5]) if len(row) > 5 else 0
        fee         = _clean(row.iloc[6]) if len(row) > 6 else 0
        misc        = _clean(row.iloc[7]) if len(row) > 7 else 0
        client_name = str(row.iloc[9]).strip() if len(row) > 9 else ""
        invoice_num = row.iloc[10]        if len(row) > 10 else None

        # Skip rows without a valid date (e.g. total row at the bottom)
        if pd.isna(txn_date) or not hasattr(txn_date, "month"):
            continue
        if not client_name or client_name in ("nan", "לקוח", ""):
            continue

        txn_date_obj = txn_date.date() if hasattr(txn_date, "date") else txn_date

        # Client lookup
        account, ratio, matched_name = flexible_match(client_name, clients_dict)
        if account is None:
            if client_name not in unmatched:
                unmatched.append(client_name)
            continue

        # Build column → value map for non-zero columns
        col_data = {}
        if fee      > 0: col_data["שכט"]     = fee
        if notary   > 0: col_data["נוטריון"] = notary
        if expenses > 0: col_data["הוצאות"]  = expenses
        if misc     > 0: col_data["שונות"]   = misc

        # Detect unknown columns
        for col_name in col_data:
            if col_name not in KNOWN_COLUMNS and col_name not in extra_columns:
                unknown_columns.add(col_name)

        # Skip row if any column is still unknown
        if any(c not in KNOWN_COLUMNS and c not in extra_columns for c in col_data):
            continue

        # ── Invoice rows ──
        for col_name, col_value in col_data.items():
            if col_name in KNOWN_COLUMNS:
                credit_account, is_vat_exempt = KNOWN_COLUMNS[col_name]
            else:
                cfg            = extra_columns[col_name]
                credit_account = cfg["account"]
                is_vat_exempt  = cfg["vat_exempt"]

            if is_vat_exempt:
                row_total  = round(col_value, 2)
                row_income = round(col_value, 2)
                row_vat    = 0.0
            else:
                row_vat    = round(col_value * VAT_RATE, 2)
                row_income = round(col_value, 2)
                row_total  = round(col_value + row_vat, 2)

            invoice_rows.append([
                txn_date_obj,
                account, credit_account, 9001,
                client_name, invoice_num,
                row_total, row_income, row_vat,
            ])

        # ── Receipt row (one per invoice, full total) ──
        if total > 0:
            receipt_rows.append([
                txn_date_obj,
                1200, account,
                client_name, invoice_num,
                total, total,
            ])

    return invoice_rows, receipt_rows, unmatched, list(unknown_columns)


def create_excel_output(invoice_rows: list, receipt_rows: list, month_label: str) -> io.BytesIO:
    """Build Excel workbook with two sheets: invoices and receipts."""
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)

    def write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append(row)
        # Date format on column A
        for r in ws.iter_rows(min_row=2, max_col=1):
            for cell in r:
                cell.number_format = "DD/MM/YYYY"
        # Auto column width
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max(12, width + 2)

    ws_inv = wb.active
    ws_inv.title = month_label
    write_sheet(ws_inv, INVOICE_HEADERS, invoice_rows)

    ws_rec = wb.create_sheet(f"{month_label} קבלות")
    write_sheet(ws_rec, RECEIPT_HEADERS, receipt_rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
